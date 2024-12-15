import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.drawing.image import Image
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.drawing.image import Image
import matplotlib.pyplot as plt
import matplotlib.pyplot as plt
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, SessionNotCreatedException
import time
import win32com.client
import tkinter as tk
from tkinter import messagebox

try:
    # Initialize WebDriver service
    service = Service(
        executable_path="chromedriver.exe"
    )
    driver = webdriver.Chrome(service=service)
    driver.get("https://slate.uaonline.arizona.edu/manage/")
    time.sleep(2)
except SessionNotCreatedException:
    print("The ChromeDriver needs to be updated. Please visit the following page to download the latest version: https://chromedriver.chromium.org/downloads")
    messagebox.showinfo("Notification", "The ChromeDriver needs to be updated. Please visit the following page to download the latest version: https://chromedriver.chromium.org/downloads")

# Open target URL
driver.get("https://slate.uaonline.arizona.edu/manage/")
time.sleep(2)

# Wait for the page to load
WebDriverWait(driver, 600).until(EC.presence_of_element_located((By.ID, "qs_suggest")))

# Target report URL and data structure
url = "https://slate.uaonline.arizona.edu/manage/report/render?id=5fc2c1f3-e030-498d-8e23-fd9892118a8d"

xpaths = [
    '//*[@id="report_part_e573da7a-4ef2-4882-bd2c-34c1f45c1973"]/div[2]/table/tbody',
    '//*[@id="report_part_cd827541-acc9-4e12-95d6-6740e0aa712d"]/div[2]/table/tbody',
    '//*[@id="report_part_ebdc6ad3-38f4-449e-8707-1fd665bb2acb"]/div[2]/table/tbody',
    '//*[@id="report_part_a20e3026-f75c-49fd-b1ed-1ff08889e88f"]/div[2]/table/tbody',
    '//*[@id="report_part_2cbd10b9-fb26-4795-a774-acf2ccfa6b25"]/div[2]/table/tbody',
    '//*[@id="report_part_c76d9529-18bf-42ef-a7a9-1754f49731bb"]/div[2]/table/tbody'
]

labels = [
    "Partial",
    "Application Entered",
    "Application Under Review",
    "Application Admitted",
    "Application Matric",
    "Application Enrolled"
]

columns = [
    ['Student ID', 'First', 'Last', 'UA Academic', 'App Created', 'App Submitted', 'Report Status'],
    ['Student ID', 'First', 'Last', 'UA Academic', 'App Created', 'App Submitted', 'Report Status'],
    ['Student ID', 'First', 'Last', 'UA Academic', 'App Created', 'App Submitted', 'Application Status Date',
     'Report Status'],
    ['Student ID', 'First', 'Last', 'UA Academic', 'Student Group', 'App Submitted', 'Admitted Date', 'PM Date',
     'Report Status'],
    ['Student ID', 'First', 'Last', 'UA Academic', 'Student Group', 'App Submitted', 'Admitted Date', 'PM Date',
     'Matric Date', 'Report Status'],
    ['Student ID', 'First', 'Last', 'UA Academic', 'App Submitted', 'Matric Date', 'Enrolled Date', 'Report Status']
]

# Extract data from the page Slate
all_data = []
driver.get(url)
time.sleep(3)

for xpath, label, cols in zip(xpaths, labels, columns):
    time.sleep(1)
    tbody = driver.find_element(By.XPATH, xpath)
    data = [[col.text for col in tr.find_elements(By.XPATH, './td')] + [label] for tr in
            tbody.find_elements(By.XPATH, './tr')]
    all_data.append(pd.DataFrame(data, columns=cols))

df = pd.concat(all_data, ignore_index=True)
print(df)

# Base URL for UAccess
base_url = "https://student851.uaccess.arizona.edu/psp/uazsaprd/EMPLOYEE/HRMS/c/UA_SA_CUSTOM.UA_AD010_APP_SUMRY.GBL"

students_data = []
students_manual = []

driver.get(base_url)
time.sleep(3)

for _, row in df.iterrows():
    student_id = row['Student ID']

    driver.get(base_url)

    if student_id.strip() == "":
        print(f"ID N/A: {student_id}")
        student_info = {
            "Student ID": student_id,
            "First": row['First'],
            "Last": row['Last'],
            "UA Academic": row['UA Academic'],
            "Report Status": row['Report Status']
        }
        students_data.append(student_info)
        continue

    try:
        WebDriverWait(driver, 30).until(EC.frame_to_be_available_and_switch_to_it(0))

        id_input = WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "UA_ADM_SUM_SCTY_EMPLID")))
        id_input.clear()
        id_input.send_keys(student_id)

        search_button = WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((By.ID, "PTS_CFG_CL_WRK_PTS_SRCH_BTN")))
        search_button.click()

        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, 'UA_APSM_BIODM_D_ADDRESSLONG')))

        student_info = {
            "Student ID": student_id,
            "First": row['First'],
            "Last": row['Last'],
            "Primary Academic Plan": "",
            "App Created": row['App Created'],
            "App Submitted": row['App Submitted'],
            "Application Status Date": row['Application Status Date'],
            "Admitted Date": row['Admitted Date'],
            "Matric Date": row["Matric Date"],
            "Enrolled Date": row['Enrolled Date'],
            "Term": "",
            "College Name": "",
            "Status": "",
            "Report Status": row['Report Status'],
            "Application Fee Status": "",
            "Aplication Fee": "",
            "College Trancript": "",
            "High School Trancript": "",
            "College Transcript": "",
            "Financial Aid Eligibility": "",
            "Email": "",
            "Phone": "",
            "Address": "",
            "Notes": ""
        }

        # Attempt to find additional fields
        fields = [
            ("Primary Academic Plan", 'ACAD_PLAN_TBL_DESCR$0'),
            ("Term", 'TERM_TBL_DESCR$0'),
            ("College Name", 'UA_PERS_CHKITEM_NAME$1'),
            ("Status", 'UA_PERS_CHKITEM_ITEM_STATUS$1'),
            ("Application Fee Status", 'UA_PERS_CHKITEM_ITEM_STATUS$0'),
            ("Aplication Fee", 'UA_PERS_CHKITEM_ITEM_STATUS$0'),
            ("College Trancript", 'UA_PERS_CHKITEM_ITEM_STATUS$1'),
            ("High School Trancript", 'UA_PERS_CHKITEM_ITEM_STATUS$2'),
            ("College Transcript", 'UA_PERS_CHKITEM_ITEM_STATUS$3'),
            ("Financial Aid Eligibility", 'UA_PERS_CHKITEM_ITEM_STATUS$4'),
            ("Email", 'UA_APSM_BIODM_D_EMAIL_ADDR'),
            ("Phone", 'UA_APSM_BIODM_D_PHONE_DISPLAY'),
            ("Address", 'UA_APSM_BIODM_D_ADDRESSLONG'),
            ("Notes", 'PERSON_COMMENT_COMMENTS$0'),
        ]

        for field_name, field_id in fields:
            try:
                student_info[field_name] = driver.find_element(By.ID, field_id).text
            except:
                pass

        students_data.append(student_info)
    except Exception as e:
        print(f"Error processing student {student_id}: {str(e)}")
        print(f"Student ID {student_id} must be registered manually.")
        students_manual.append(student_id)

# Path to the Excel file
file_path = 'students_info.xlsx'

try:
    # Read the Excel file
    df = pd.read_excel(file_path)

    # Print available columns
    print("Columns in the file:", df.columns)

    # Verify and process Academic Plan columns
    if 'UA Academic' in df.columns and 'Primary Academic Plan' in df.columns:
        # Replace NaN values with empty strings
        df['UA Academic'] = df['UA Academic'].fillna('')
        df['Primary Academic Plan'] = df['Primary Academic Plan'].fillna('')

        # Combine columns
        df['Primary Academic Plan'] = df['UA Academic'] + ' ' + df['Primary Academic Plan']
    else:
        print("Warning: Columns 'UA Academic' or 'Primary Academic Plan' not found")

    # Verify and process Report Status column
    if 'Report Status' in df.columns:
        # Retrieve unique values of Report Status
        status_categories = df['Report Status'].unique()

        # Load the workbook with openpyxl to handle multiple sheets
        wb = load_workbook(file_path)

        # Create sheets for each Report Status category
        for status in status_categories:
            # Filter the dataframe for the current category
            filtered_df = df[df['Report Status'] == status]

            # Create a valid sheet name (replace invalid characters)
            sheet_name = str(status).replace('/', '_').replace('\\', '_')[:31]

            # Remove the sheet if it already exists
            if sheet_name in wb.sheetnames:
                wb.remove(wb[sheet_name])

            # Create a new sheet
            ws = wb.create_sheet(sheet_name)

            # Write headers and data in a single operation
            for r_idx, row in enumerate(filtered_df.values.tolist(), start=1):
                for c_idx, value in enumerate(filtered_df.columns, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=value if r_idx == 1 else row[c_idx - 1])

        print(f"Categories found: {list(status_categories)}")

        # Apply styles to all sheets
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Iterate through the cells and apply styles
            for row in ws.iter_rows(min_row=1, min_col=1, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    if cell.value in ["Waived", "Completed"]:
                        cell.font = Font(color="FFFFFF")  # White text (invisible on white background)
                    elif cell.value == "Required":
                        cell.font = Font(bold=True)  # Bold text

        # Save the changes
        wb.save(file_path)
        print(f"File processed and saved at: {file_path}")

except Exception as e:
    print(f"An error occurred: {e}")

# File configuration
input_file = 'students_info.xlsx'
output_file = 'students_info_cleaned.xlsx'

# Dictionary mapping each sheet with columns to remove
columns_to_remove = {
    'Partial': ['App Submitted', 'Application Status Date', 'Admitted Date', 'Matric Date', 'Enrolled Date'],
    'Application Entered': ['App Created', 'Application Status Date', 'Admitted Date', 'Matric Date', 'Enrolled Date'],
    'Application Under Review': ['App Created', 'App Submitted', 'Admitted Date', 'Matric Date', 'Enrolled Date'],
    'Application Admitted': ['App Created', 'App Submitted', 'Application Status Date', 'Matric Date', 'Enrolled Date'],
    'Application Matric': ['App Created', 'App Submitted', 'Application Status Date', 'Admitted Date', 'Enrolled Date'],
    'Application Enrolled': ['App Created', 'App Submitted', 'Application Status Date', 'Admitted Date', 'Matric Date']
}

# Open the Excel file
excel_file = pd.ExcelFile(input_file)

# Create an Excel writer to save the changes
with pd.ExcelWriter(output_file) as writer:
    # Iterate over each sheet in the file
    for sheet_name in excel_file.sheet_names:
        # Read the DataFrame for each sheet
        df = pd.read_excel(input_file, sheet_name=sheet_name)

        # Remove the specified columns if they exist in the sheet
        if sheet_name in columns_to_remove:
            columns_to_drop = [col for col in columns_to_remove[sheet_name] if col in df.columns]
            df = df.drop(columns=columns_to_drop)

        # Save the modified DataFrame to a new sheet in the output file
        df.to_excel(writer, sheet_name=sheet_name, index=False)

print(f"File processed. Removed columns saved in {output_file}")

# Read Excel file
summary_df = pd.read_excel("students_info.xlsx")

# Summary by category
ua_academic_summary = summary_df['Primary Academic Plan'].value_counts()
report_status_summary = summary_df['Report Status'].value_counts()
total_students = len(summary_df)

# Create Excel file with summary
with pd.ExcelWriter('student_summary.xlsx', engine='openpyxl') as writer:
    ua_academic_summary.to_excel(writer, sheet_name='Primary Academic Plan')
    report_status_summary.to_excel(writer, sheet_name='Report_Status')
    pd.DataFrame({'Total Students': [total_students]}).to_excel(writer, sheet_name='Total')

    workbook = writer.book

    # Plot and add to Excel
    plt.figure(figsize=(10, 6))
    ua_academic_summary.plot(kind='bar')
    plt.title('Students by UA Academic Category')
    plt.xlabel('UA Academic Category')
    plt.ylabel('Number of Students')
    plt.xticks(rotation=45, ha='right')
    plt.tight_layout()
    plt.savefig('ua_academic_chart.png')
    plt.close()

    ws_ua = workbook['Primary Academic Plan']
    img_ua = Image('ua_academic_chart.png')
    img_ua.width = 600
    img_ua.height = 400
    ws_ua.add_image(img_ua, 'E2')

    plt.figure(figsize=(10, 6))
    report_status_summary.plot(kind='bar')
    plt.title('Students by Report Status')
    plt.xlabel('Report Status')
    plt.ylabel('Number of Students')
    plt.xticks(rotation=45, ha='right')
    plt.tight_layout()
    plt.savefig('report_status_chart.png')
    plt.close()

    ws_status = workbook['Report_Status']
    img_status = Image('report_status_chart.png')
    img_status.width = 600
    img_status.height = 400
    ws_status.add_image(img_status, 'E2')

print("Analysis completed. Check 'student_summary.xlsx'.")
messagebox.showinfo("Notification",
                    "Analysis completed. Check 'student_summary.xlsx' 'student_info_cleaned.xlsx' and 'student_info.xlsx' ")